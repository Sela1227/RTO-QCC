"""報表服務"""
from typing import List, Dict
from datetime import date, timedelta
from app.models.database import Database
from app.config.constants import CANCER_TYPES, TREATMENT_STATUS, SKIP_REASONS, get_label


class ReportService:
    """報表服務"""
    
    def __init__(self):
        self.db = Database()
    
    def get_summary(self, start_date: date, end_date: date, filters: dict = None) -> dict:
        """取得統計摘要（支援篩選）"""
        filters = filters or {}
        start_str = start_date.isoformat()
        end_str = end_date.isoformat()
        
        # 建立篩選條件
        where_clauses = ["date(t.created_at) BETWEEN ? AND ?"]
        params = [start_str, end_str]
        
        if filters.get("cancer_type"):
            where_clauses.append("t.cancer_type = ?")
            params.append(filters["cancer_type"])
        
        if filters.get("exclude_unable"):
            where_clauses.append("(t.unable_to_measure = 0 OR t.unable_to_measure IS NULL)")
        
        if filters.get("status"):
            where_clauses.append("t.status = ?")
            params.append(filters["status"])
        
        where_sql = " AND ".join(where_clauses)
        
        # 基本統計
        new_patients = self.db.fetch_one(
            f"""SELECT COUNT(DISTINCT t.patient_id) as cnt FROM treatments t
               WHERE {where_sql}""", tuple(params)
        )
        
        new_treatments = self.db.fetch_one(
            f"""SELECT COUNT(*) as cnt FROM treatments t
               WHERE {where_sql}""", tuple(params)
        )
        
        # 體重記錄數
        weight_records = self.db.fetch_one(
            f"""SELECT COUNT(*) as cnt FROM weight_records w
               JOIN treatments t ON w.treatment_id = t.id
               WHERE {where_sql} AND date(w.measure_date) BETWEEN ? AND ?""",
            tuple(params) + (start_str, end_str)
        )
        
        # SDM 完成數
        sdm_completed = self.db.fetch_one(
            f"""SELECT COUNT(*) as cnt FROM interventions i
               JOIN treatments t ON i.treatment_id = t.id
               WHERE i.type='sdm' AND i.status='completed' AND {where_sql}""", tuple(params)
        )
        
        # 營養師完成數
        nutrition_completed = self.db.fetch_one(
            f"""SELECT COUNT(*) as cnt FROM interventions i
               JOIN treatments t ON i.treatment_id = t.id
               WHERE i.type='nutrition' AND i.status='completed' AND {where_sql}""", tuple(params)
        )
        
        # 結案數
        completed = self.db.fetch_one(
            f"""SELECT COUNT(*) as cnt FROM treatments t
               WHERE t.status='completed' AND {where_sql}""", tuple(params)
        )
        
        return {
            "new_patients": new_patients["cnt"] if new_patients else 0,
            "new_treatments": new_treatments["cnt"] if new_treatments else 0,
            "weight_records": weight_records["cnt"] if weight_records else 0,
            "sdm_completed": sdm_completed["cnt"] if sdm_completed else 0,
            "nutrition_completed": nutrition_completed["cnt"] if nutrition_completed else 0,
            "completed_treatments": completed["cnt"] if completed else 0,
        }
    
    def get_weight_distribution(self, start_date: date, end_date: date, filters: dict = None) -> dict:
        """取得體重變化分布"""
        treatments = self.get_treatment_report(start_date, end_date, filters)
        
        severe_loss = 0  # ≤-5%
        moderate_loss = 0  # -5% < x ≤ -3%
        maintained = 0  # > -3%
        
        for t in treatments:
            rate = t["change_rate"] or 0
            if rate <= -5:
                severe_loss += 1
            elif rate <= -3:
                moderate_loss += 1
            else:
                maintained += 1
        
        total = len(treatments)
        return {
            "severe_loss": severe_loss,
            "severe_loss_pct": round(severe_loss / total * 100, 1) if total > 0 else 0,
            "moderate_loss": moderate_loss,
            "moderate_loss_pct": round(moderate_loss / total * 100, 1) if total > 0 else 0,
            "maintained": maintained,
            "maintained_pct": round(maintained / total * 100, 1) if total > 0 else 0,
            "total": total,
        }
    
    def get_intervention_stats(self, start_date: date, end_date: date, filters: dict = None) -> dict:
        """取得介入執行統計"""
        filters = filters or {}
        start_str = start_date.isoformat()
        end_str = end_date.isoformat()
        
        where_clauses = ["date(i.created_at) BETWEEN ? AND ?"]
        params = [start_str, end_str]
        
        if filters.get("cancer_type"):
            where_clauses.append("t.cancer_type = ?")
            params.append(filters["cancer_type"])
        
        if filters.get("exclude_unable"):
            where_clauses.append("(t.unable_to_measure = 0 OR t.unable_to_measure IS NULL)")
        
        where_sql = " AND ".join(where_clauses)
        
        # SDM 統計
        sdm_stats = self.db.fetch_one(
            f"""SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN i.status='completed' THEN 1 ELSE 0 END) as completed,
                SUM(CASE WHEN i.status='skipped' THEN 1 ELSE 0 END) as skipped
               FROM interventions i
               JOIN treatments t ON i.treatment_id = t.id
               WHERE i.type='sdm' AND {where_sql}""", tuple(params)
        )
        
        # 營養師統計
        nutrition_stats = self.db.fetch_one(
            f"""SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN i.status='completed' THEN 1 ELSE 0 END) as completed,
                SUM(CASE WHEN i.status='skipped' THEN 1 ELSE 0 END) as skipped
               FROM interventions i
               JOIN treatments t ON i.treatment_id = t.id
               WHERE i.type='nutrition' AND {where_sql}""", tuple(params)
        )
        
        # 略過原因分布
        skip_reasons = self.db.fetch_all(
            f"""SELECT skip_reason, COUNT(*) as cnt
               FROM interventions i
               JOIN treatments t ON i.treatment_id = t.id
               WHERE i.status='skipped' AND {where_sql}
               GROUP BY skip_reason""", tuple(params)
        )
        
        skip_distribution = {}
        for row in skip_reasons:
            reason = row["skip_reason"] or "OTHER"
            skip_distribution[reason] = row["cnt"]
        
        sdm_total = sdm_stats["total"] if sdm_stats else 0
        sdm_completed = sdm_stats["completed"] if sdm_stats else 0
        nutrition_total = nutrition_stats["total"] if nutrition_stats else 0
        nutrition_completed = nutrition_stats["completed"] if nutrition_stats else 0
        
        return {
            "sdm": {
                "total": sdm_total,
                "completed": sdm_completed,
                "skipped": sdm_stats["skipped"] if sdm_stats else 0,
                "rate": round(sdm_completed / sdm_total * 100, 1) if sdm_total > 0 else 0,
            },
            "nutrition": {
                "total": nutrition_total,
                "completed": nutrition_completed,
                "skipped": nutrition_stats["skipped"] if nutrition_stats else 0,
                "rate": round(nutrition_completed / nutrition_total * 100, 1) if nutrition_total > 0 else 0,
            },
            "skip_distribution": skip_distribution,
        }
    
    def get_intervention_effectiveness(self, start_date: date, end_date: date, filters: dict = None) -> dict:
        """取得介入成效統計（介入後體重回升比率）"""
        filters = filters or {}
        start_str = start_date.isoformat()
        end_str = end_date.isoformat()
        
        where_clauses = ["date(i.completed_at) BETWEEN ? AND ?", "i.status = 'completed'"]
        params = [start_str, end_str]
        
        if filters.get("cancer_type"):
            where_clauses.append("t.cancer_type = ?")
            params.append(filters["cancer_type"])
        
        where_sql = " AND ".join(where_clauses)
        
        # 取得已完成的介入及其關聯的體重記錄
        interventions = self.db.fetch_all(
            f"""SELECT i.id, i.type, i.treatment_id, i.completed_at,
                       t.baseline_weight
               FROM interventions i
               JOIN treatments t ON i.treatment_id = t.id
               WHERE {where_sql}
               ORDER BY i.completed_at""", tuple(params)
        )
        
        sdm_improved = 0
        sdm_total = 0
        nutrition_improved = 0
        nutrition_total = 0
        
        for inv in interventions:
            treatment_id = inv["treatment_id"]
            completed_at = inv["completed_at"]
            baseline = inv["baseline_weight"]
            inv_type = inv["type"]
            
            if not completed_at or not baseline:
                continue
            
            # 取得介入完成時的體重（最接近完成日期的記錄）
            weight_at_intervention = self.db.fetch_one(
                """SELECT weight FROM weight_records 
                   WHERE treatment_id = ? AND measure_date <= ?
                   ORDER BY measure_date DESC LIMIT 1""",
                (treatment_id, completed_at.split('T')[0] if 'T' in str(completed_at) else completed_at)
            )
            
            # 取得介入後 14 天內的最新體重
            weight_after = self.db.fetch_one(
                """SELECT weight FROM weight_records 
                   WHERE treatment_id = ? AND measure_date > ?
                   ORDER BY measure_date ASC LIMIT 1""",
                (treatment_id, completed_at.split('T')[0] if 'T' in str(completed_at) else completed_at)
            )
            
            if weight_at_intervention and weight_after:
                before_weight = weight_at_intervention["weight"]
                after_weight = weight_after["weight"]
                
                # 判斷是否改善（體重回升或變化率改善）
                before_rate = (before_weight - baseline) / baseline * 100
                after_rate = (after_weight - baseline) / baseline * 100
                
                # 改善定義：體重回升超過0.5kg 或 變化率改善超過1%
                improved = (after_weight > before_weight + 0.5) or (after_rate > before_rate + 1)
                
                if inv_type == "sdm":
                    sdm_total += 1
                    if improved:
                        sdm_improved += 1
                elif inv_type == "nutrition":
                    nutrition_total += 1
                    if improved:
                        nutrition_improved += 1
        
        return {
            "sdm": {
                "total": sdm_total,
                "improved": sdm_improved,
                "rate": round(sdm_improved / sdm_total * 100, 1) if sdm_total > 0 else 0,
            },
            "nutrition": {
                "total": nutrition_total,
                "improved": nutrition_improved,
                "rate": round(nutrition_improved / nutrition_total * 100, 1) if nutrition_total > 0 else 0,
            },
            "overall": {
                "total": sdm_total + nutrition_total,
                "improved": sdm_improved + nutrition_improved,
                "rate": round((sdm_improved + nutrition_improved) / (sdm_total + nutrition_total) * 100, 1) if (sdm_total + nutrition_total) > 0 else 0,
            }
        }
    
    def get_tracking_quality(self, start_date: date, end_date: date, filters: dict = None) -> dict:
        """取得追蹤品質統計"""
        treatments = self.get_treatment_report(start_date, end_date, filters)
        
        if not treatments:
            return {"on_time_rate": 0, "avg_interval": 0, "total": 0}
        
        intervals = []
        on_time_count = 0
        
        for t in treatments:
            if t["weight_count"] > 1:
                records = self.db.fetch_all(
                    """SELECT measure_date FROM weight_records 
                       WHERE treatment_id = ? ORDER BY measure_date""",
                    (t["id"],)
                )
                
                if len(records) >= 2:
                    for i in range(1, len(records)):
                        try:
                            d1 = date.fromisoformat(str(records[i-1]["measure_date"])[:10])
                            d2 = date.fromisoformat(str(records[i]["measure_date"])[:10])
                            interval = (d2 - d1).days
                            intervals.append(interval)
                            if interval <= 7:
                                on_time_count += 1
                        except:
                            pass
        
        total_intervals = len(intervals)
        avg_interval = round(sum(intervals) / total_intervals, 1) if total_intervals > 0 else 0
        on_time_rate = round(on_time_count / total_intervals * 100, 1) if total_intervals > 0 else 0
        
        return {
            "on_time_rate": on_time_rate,
            "avg_interval": avg_interval,
            "total_measurements": total_intervals,
        }
    
    def get_cancer_analysis(self, start_date: date, end_date: date, filters: dict = None) -> List[dict]:
        """取得癌別分析"""
        filters = filters or {}
        start_str = start_date.isoformat()
        end_str = end_date.isoformat()
        
        where_clauses = ["date(t.created_at) BETWEEN ? AND ?"]
        params = [start_str, end_str]
        
        if filters.get("exclude_unable"):
            where_clauses.append("(t.unable_to_measure = 0 OR t.unable_to_measure IS NULL)")
        
        if filters.get("status"):
            where_clauses.append("t.status = ?")
            params.append(filters["status"])
        
        where_sql = " AND ".join(where_clauses)
        
        results = []
        for ct in CANCER_TYPES:
            cancer_type = ct["code"]
            
            count = self.db.fetch_one(
                f"""SELECT COUNT(*) as cnt FROM treatments t
                   WHERE t.cancer_type = ? AND {where_sql}""",
                (cancer_type,) + tuple(params)
            )
            
            if count and count["cnt"] > 0:
                avg_rate = self.db.fetch_one(
                    f"""SELECT AVG(
                        (w.weight - t.baseline_weight) / t.baseline_weight * 100
                    ) as avg_rate
                    FROM treatments t
                    JOIN (
                        SELECT treatment_id, weight, 
                               ROW_NUMBER() OVER (PARTITION BY treatment_id ORDER BY measure_date DESC) as rn
                        FROM weight_records
                    ) w ON w.treatment_id = t.id AND w.rn = 1
                    WHERE t.cancer_type = ? AND {where_sql}""",
                    (cancer_type,) + tuple(params)
                )
                
                intervention_count = self.db.fetch_one(
                    f"""SELECT COUNT(DISTINCT i.treatment_id) as cnt
                       FROM interventions i
                       JOIN treatments t ON i.treatment_id = t.id
                       WHERE t.cancer_type = ? AND {where_sql}""",
                    (cancer_type,) + tuple(params)
                )
                
                total_count = count["cnt"]
                int_count = intervention_count["cnt"] if intervention_count else 0
                
                results.append({
                    "code": cancer_type,
                    "label": ct["label"],
                    "count": total_count,
                    "avg_change_rate": round(avg_rate["avg_rate"], 1) if avg_rate and avg_rate["avg_rate"] else 0,
                    "intervention_rate": round(int_count / total_count * 100, 1) if total_count > 0 else 0,
                })
        
        results.sort(key=lambda x: x["count"], reverse=True)
        return results
    
    def get_treatment_report(self, start_date: date, end_date: date, filters: dict = None) -> List[dict]:
        """取得療程報表（支援篩選）"""
        filters = filters or {}
        start_str = start_date.isoformat()
        end_str = end_date.isoformat()
        
        where_clauses = ["date(t.created_at) BETWEEN ? AND ?"]
        params = [start_str, end_str]
        
        if filters.get("cancer_type"):
            where_clauses.append("t.cancer_type = ?")
            params.append(filters["cancer_type"])
        
        if filters.get("exclude_unable"):
            where_clauses.append("(t.unable_to_measure = 0 OR t.unable_to_measure IS NULL)")
        
        if filters.get("status"):
            where_clauses.append("t.status = ?")
            params.append(filters["status"])
        
        where_sql = " AND ".join(where_clauses)
        
        rows = self.db.fetch_all(
            f"""SELECT 
                t.id, t.treatment_name, t.cancer_type, t.treatment_start, 
                t.treatment_end, t.baseline_weight, t.status,
                t.unable_to_measure, t.unable_reason,
                p.medical_id, p.name, p.gender,
                (SELECT COUNT(*) FROM weight_records WHERE treatment_id = t.id) as weight_count,
                (SELECT weight FROM weight_records WHERE treatment_id = t.id 
                 ORDER BY measure_date DESC LIMIT 1) as last_weight,
                (SELECT measure_date FROM weight_records WHERE treatment_id = t.id 
                 ORDER BY measure_date DESC LIMIT 1) as last_measure_date
               FROM treatments t
               JOIN patients p ON t.patient_id = p.id
               WHERE {where_sql}
               ORDER BY t.created_at DESC""",
            tuple(params)
        )
        
        result = []
        for row in rows:
            change_rate = 0
            if row["baseline_weight"] and row["last_weight"]:
                change_rate = round(
                    (row["last_weight"] - row["baseline_weight"]) / row["baseline_weight"] * 100, 
                    1
                )
            
            result.append({
                "id": row["id"],
                "medical_id": row["medical_id"],
                "name": row["name"],
                "gender": row["gender"],
                "cancer_type": row["cancer_type"],
                "treatment_name": row["treatment_name"],
                "treatment_start": row["treatment_start"],
                "treatment_end": row["treatment_end"],
                "baseline_weight": row["baseline_weight"],
                "last_weight": row["last_weight"],
                "change_rate": change_rate,
                "weight_count": row["weight_count"],
                "last_measure_date": row["last_measure_date"],
                "status": row["status"],
                "unable_to_measure": row["unable_to_measure"],
                "unable_reason": row["unable_reason"],
            })
        
        return result
    
    def get_intervention_report(self, start_date: date, end_date: date, filters: dict = None) -> List[dict]:
        """取得介入報表"""
        filters = filters or {}
        start_str = start_date.isoformat()
        end_str = end_date.isoformat()
        
        where_clauses = ["date(i.created_at) BETWEEN ? AND ?"]
        params = [start_str, end_str]
        
        if filters.get("cancer_type"):
            where_clauses.append("t.cancer_type = ?")
            params.append(filters["cancer_type"])
        
        where_sql = " AND ".join(where_clauses)
        
        rows = self.db.fetch_all(
            f"""SELECT 
                i.id, i.type, i.source, i.status, i.trigger_reason,
                i.executed_at, i.skip_reason, i.created_at,
                t.treatment_name, t.cancer_type,
                p.medical_id, p.name
               FROM interventions i
               JOIN treatments t ON i.treatment_id = t.id
               JOIN patients p ON t.patient_id = p.id
               WHERE {where_sql}
               ORDER BY i.created_at DESC""",
            tuple(params)
        )
        
        return [dict(row) for row in rows]
    
    def export_to_excel(self, start_date: date, end_date: date, filepath: str, filters: dict = None) -> bool:
        """匯出報表到 Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            
            # === 統計摘要表 ===
            ws = wb.active
            ws.title = "統計摘要"
            
            summary = self.get_summary(start_date, end_date, filters)
            weight_dist = self.get_weight_distribution(start_date, end_date, filters)
            int_stats = self.get_intervention_stats(start_date, end_date, filters)
            tracking = self.get_tracking_quality(start_date, end_date, filters)
            
            ws["A1"] = "放射線治療體重追蹤系統 - 統計報表"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A2"] = f"期間：{start_date} ~ {end_date}"
            
            filter_text = []
            if filters:
                if filters.get("cancer_type"):
                    filter_text.append(f"癌別：{get_label(CANCER_TYPES, filters['cancer_type'])}")
                if filters.get("exclude_unable"):
                    filter_text.append("排除無法測量")
                if filters.get("status"):
                    filter_text.append(f"狀態：{get_label(TREATMENT_STATUS, filters['status'])}")
            ws["A3"] = f"篩選：{', '.join(filter_text)}" if filter_text else ""
            
            row = 5
            ws[f"A{row}"] = "【基本統計】"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
            for label, value in [
                ("新增病人數", summary["new_patients"]),
                ("新增療程數", summary["new_treatments"]),
                ("體重記錄數", summary["weight_records"]),
                ("結案療程數", summary["completed_treatments"]),
            ]:
                ws[f"A{row}"] = label
                ws[f"B{row}"] = value
                row += 1
            
            row += 1
            ws[f"A{row}"] = "【體重變化分布】"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
            ws[f"A{row}"] = "下降 ≥5%（需營養師）"
            ws[f"B{row}"] = f"{weight_dist['severe_loss']} 人 ({weight_dist['severe_loss_pct']}%)"
            row += 1
            ws[f"A{row}"] = "下降 3-5%（需 SDM）"
            ws[f"B{row}"] = f"{weight_dist['moderate_loss']} 人 ({weight_dist['moderate_loss_pct']}%)"
            row += 1
            ws[f"A{row}"] = "維持/上升"
            ws[f"B{row}"] = f"{weight_dist['maintained']} 人 ({weight_dist['maintained_pct']}%)"
            
            row += 2
            ws[f"A{row}"] = "【介入執行率】"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
            ws[f"A{row}"] = "SDM 執行率"
            ws[f"B{row}"] = f"{int_stats['sdm']['rate']}% ({int_stats['sdm']['completed']}/{int_stats['sdm']['total']})"
            row += 1
            ws[f"A{row}"] = "營養師轉介執行率"
            ws[f"B{row}"] = f"{int_stats['nutrition']['rate']}% ({int_stats['nutrition']['completed']}/{int_stats['nutrition']['total']})"
            
            row += 2
            ws[f"A{row}"] = "【追蹤品質】"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
            ws[f"A{row}"] = "按時量測率（7天內）"
            ws[f"B{row}"] = f"{tracking['on_time_rate']}%"
            row += 1
            ws[f"A{row}"] = "平均量測間隔"
            ws[f"B{row}"] = f"{tracking['avg_interval']} 天"
            
            if int_stats["skip_distribution"]:
                row += 2
                ws[f"A{row}"] = "【略過原因分布】"
                ws[f"A{row}"].font = Font(bold=True)
                row += 1
                for reason, cnt in int_stats["skip_distribution"].items():
                    ws[f"A{row}"] = get_label(SKIP_REASONS, reason)
                    ws[f"B{row}"] = cnt
                    row += 1
            
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 20
            
            # === 癌別分析表 ===
            ws_cancer = wb.create_sheet("癌別分析")
            cancer_data = self.get_cancer_analysis(start_date, end_date, filters)
            
            headers = ["癌別", "療程數", "平均變化率(%)", "介入率(%)"]
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, start=1):
                cell = ws_cancer.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            for row_idx, cd in enumerate(cancer_data, start=2):
                ws_cancer.cell(row=row_idx, column=1, value=cd["label"])
                ws_cancer.cell(row=row_idx, column=2, value=cd["count"])
                ws_cancer.cell(row=row_idx, column=3, value=cd["avg_change_rate"])
                ws_cancer.cell(row=row_idx, column=4, value=cd["intervention_rate"])
            
            for col in range(1, 5):
                ws_cancer.column_dimensions[chr(64 + col)].width = 15
            
            # === 療程清單表 ===
            ws_treatment = wb.create_sheet("療程清單")
            t_headers = ["病歷號", "姓名", "癌別", "基準體重", "目前體重", "變化率(%)", "量測次數", "狀態"]
            
            for col, header in enumerate(t_headers, start=1):
                cell = ws_treatment.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            treatments = self.get_treatment_report(start_date, end_date, filters)
            for row_idx, t in enumerate(treatments, start=2):
                ws_treatment.cell(row=row_idx, column=1, value=t["medical_id"])
                ws_treatment.cell(row=row_idx, column=2, value=t["name"])
                ws_treatment.cell(row=row_idx, column=3, value=get_label(CANCER_TYPES, t["cancer_type"]))
                ws_treatment.cell(row=row_idx, column=4, value=t["baseline_weight"])
                ws_treatment.cell(row=row_idx, column=5, value=t["last_weight"])
                ws_treatment.cell(row=row_idx, column=6, value=t["change_rate"])
                ws_treatment.cell(row=row_idx, column=7, value=t["weight_count"])
                ws_treatment.cell(row=row_idx, column=8, value=get_label(TREATMENT_STATUS, t["status"]))
            
            for col in range(1, 9):
                ws_treatment.column_dimensions[chr(64 + col)].width = 12
            
            # === 介入記錄表 ===
            ws_int = wb.create_sheet("介入記錄")
            int_headers = ["病歷號", "姓名", "類型", "狀態", "原因", "執行日期"]
            
            for col, header in enumerate(int_headers, start=1):
                cell = ws_int.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            interventions = self.get_intervention_report(start_date, end_date, filters)
            for row_idx, i in enumerate(interventions, start=2):
                type_label = {"sdm": "SDM", "nutrition": "營養師"}.get(i["type"], i["type"])
                status_label = {"pending": "待處理", "completed": "已完成", "skipped": "已略過"}.get(i["status"], i["status"])
                
                ws_int.cell(row=row_idx, column=1, value=i["medical_id"])
                ws_int.cell(row=row_idx, column=2, value=i["name"])
                ws_int.cell(row=row_idx, column=3, value=type_label)
                ws_int.cell(row=row_idx, column=4, value=status_label)
                ws_int.cell(row=row_idx, column=5, value=i["trigger_reason"] or i["skip_reason"] or "")
                ws_int.cell(row=row_idx, column=6, value=str(i["executed_at"])[:10] if i["executed_at"] else "")
            
            for col in range(1, 7):
                ws_int.column_dimensions[chr(64 + col)].width = 12
            
            wb.save(filepath)
            return True
            
        except Exception as e:
            print(f"Excel 匯出失敗: {e}")
            return False
