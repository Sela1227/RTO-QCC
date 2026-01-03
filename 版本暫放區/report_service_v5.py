"""報表服務"""
from typing import List, Dict
from datetime import date, timedelta
from app.models.database import Database
from app.repositories.patient_repository import PatientRepository
from app.repositories.treatment_repository import TreatmentRepository
from app.repositories.weight_repository import WeightRepository
from app.repositories.intervention_repository import InterventionRepository


class ReportService:
    """報表服務"""
    
    def __init__(self):
        self.db = Database()
        self.patient_repo = PatientRepository()
        self.treatment_repo = TreatmentRepository()
        self.weight_repo = WeightRepository()
        self.intervention_repo = InterventionRepository()
    
    def get_summary(self, start_date: date = None, end_date: date = None) -> dict:
        """取得統計摘要"""
        if not start_date:
            start_date = date.today() - timedelta(days=30)
        if not end_date:
            end_date = date.today()
        
        start_str = start_date.isoformat()
        end_str = end_date.isoformat()
        
        # 期間內新增病人數
        new_patients = self.db.fetch_one(
            """SELECT COUNT(*) as cnt FROM patients 
               WHERE date(created_at) BETWEEN ? AND ?""",
            (start_str, end_str)
        )
        
        # 期間內新增療程數
        new_treatments = self.db.fetch_one(
            """SELECT COUNT(*) as cnt FROM treatments 
               WHERE date(created_at) BETWEEN ? AND ?""",
            (start_str, end_str)
        )
        
        # 期間內體重記錄數
        weight_records = self.db.fetch_one(
            """SELECT COUNT(*) as cnt FROM weight_records 
               WHERE date(measure_date) BETWEEN ? AND ?""",
            (start_str, end_str)
        )
        
        # 期間內介入數
        interventions = self.db.fetch_one(
            """SELECT COUNT(*) as cnt FROM interventions 
               WHERE date(created_at) BETWEEN ? AND ?""",
            (start_str, end_str)
        )
        
        # SDM 完成數
        sdm_completed = self.db.fetch_one(
            """SELECT COUNT(*) as cnt FROM interventions 
               WHERE type='sdm' AND status='completed' 
               AND date(executed_at) BETWEEN ? AND ?""",
            (start_str, end_str)
        )
        
        # 營養師完成數
        nutrition_completed = self.db.fetch_one(
            """SELECT COUNT(*) as cnt FROM interventions 
               WHERE type='nutrition' AND status='completed' 
               AND date(executed_at) BETWEEN ? AND ?""",
            (start_str, end_str)
        )
        
        # 結案數
        completed_treatments = self.db.fetch_one(
            """SELECT COUNT(*) as cnt FROM treatments 
               WHERE status='completed' 
               AND date(status_changed_at) BETWEEN ? AND ?""",
            (start_str, end_str)
        )
        
        return {
            "period": {"start": start_date, "end": end_date},
            "new_patients": new_patients["cnt"] if new_patients else 0,
            "new_treatments": new_treatments["cnt"] if new_treatments else 0,
            "weight_records": weight_records["cnt"] if weight_records else 0,
            "interventions": interventions["cnt"] if interventions else 0,
            "sdm_completed": sdm_completed["cnt"] if sdm_completed else 0,
            "nutrition_completed": nutrition_completed["cnt"] if nutrition_completed else 0,
            "completed_treatments": completed_treatments["cnt"] if completed_treatments else 0,
        }
    
    def get_treatment_report(self, start_date: date = None, end_date: date = None) -> List[dict]:
        """取得療程報表"""
        if not start_date:
            start_date = date.today() - timedelta(days=30)
        if not end_date:
            end_date = date.today()
        
        rows = self.db.fetch_all(
            """SELECT 
                t.id, t.treatment_name, t.cancer_type, t.treatment_start, 
                t.treatment_end, t.baseline_weight, t.status,
                p.medical_id, p.name, p.gender,
                (SELECT COUNT(*) FROM weight_records WHERE treatment_id = t.id) as weight_count,
                (SELECT weight FROM weight_records WHERE treatment_id = t.id 
                 ORDER BY measure_date DESC LIMIT 1) as last_weight,
                (SELECT measure_date FROM weight_records WHERE treatment_id = t.id 
                 ORDER BY measure_date DESC LIMIT 1) as last_measure_date
               FROM treatments t
               JOIN patients p ON t.patient_id = p.id
               WHERE date(t.created_at) BETWEEN ? AND ?
               ORDER BY t.created_at DESC""",
            (start_date.isoformat(), end_date.isoformat())
        )
        
        result = []
        for row in rows:
            # 計算體重變化（負=下降，正=上升）
            change_rate = 0
            if row["baseline_weight"] and row["last_weight"]:
                change_rate = round(
                    (row["last_weight"] - row["baseline_weight"]) / row["baseline_weight"] * 100, 
                    1
                )
            
            result.append({
                "id": row["id"],
                "medical_id": row["medical_id"],
                "name": row["name"],
                "gender": row["gender"],
                "cancer_type": row["cancer_type"],
                "treatment_name": row["treatment_name"],
                "treatment_start": row["treatment_start"],
                "treatment_end": row["treatment_end"],
                "baseline_weight": row["baseline_weight"],
                "last_weight": row["last_weight"],
                "change_rate": change_rate,
                "weight_count": row["weight_count"],
                "last_measure_date": row["last_measure_date"],
                "status": row["status"],
            })
        
        return result
    
    def get_intervention_report(self, start_date: date = None, end_date: date = None) -> List[dict]:
        """取得介入報表"""
        if not start_date:
            start_date = date.today() - timedelta(days=30)
        if not end_date:
            end_date = date.today()
        
        rows = self.db.fetch_all(
            """SELECT 
                i.id, i.type, i.source, i.status, i.trigger_reason,
                i.executed_at, i.skip_reason, i.created_at,
                t.treatment_name,
                p.medical_id, p.name
               FROM interventions i
               JOIN treatments t ON i.treatment_id = t.id
               JOIN patients p ON t.patient_id = p.id
               WHERE date(i.created_at) BETWEEN ? AND ?
               ORDER BY i.created_at DESC""",
            (start_date.isoformat(), end_date.isoformat())
        )
        
        return [dict(row) for row in rows]
    
    def export_to_excel(self, start_date: date, end_date: date, filepath: str) -> bool:
        """匯出報表到 Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            wb = Workbook()
            
            # === 摘要表 ===
            ws_summary = wb.active
            ws_summary.title = "統計摘要"
            
            summary = self.get_summary(start_date, end_date)
            
            # 標題
            ws_summary["A1"] = "放射線治療體重追蹤系統 - 統計報表"
            ws_summary["A1"].font = Font(bold=True, size=14)
            ws_summary["A2"] = f"期間：{start_date} ~ {end_date}"
            
            # 統計數據
            summary_data = [
                ("新增病人數", summary["new_patients"]),
                ("新增療程數", summary["new_treatments"]),
                ("體重記錄數", summary["weight_records"]),
                ("介入觸發數", summary["interventions"]),
                ("SDM 完成數", summary["sdm_completed"]),
                ("營養師完成數", summary["nutrition_completed"]),
                ("結案療程數", summary["completed_treatments"]),
            ]
            
            for i, (label, value) in enumerate(summary_data, start=4):
                ws_summary[f"A{i}"] = label
                ws_summary[f"B{i}"] = value
            
            # 調整欄寬
            ws_summary.column_dimensions["A"].width = 20
            ws_summary.column_dimensions["B"].width = 15
            
            # === 療程表 ===
            ws_treatment = wb.create_sheet("療程清單")
            
            headers = ["病歷號", "姓名", "性別", "癌別", "療程名稱", 
                      "開始日期", "結束日期", "基準體重", "目前體重", 
                      "變化率(%)", "量測次數", "最後量測", "狀態"]
            
            # 標題樣式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col, header in enumerate(headers, start=1):
                cell = ws_treatment.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # 資料
            treatments = self.get_treatment_report(start_date, end_date)
            for row_idx, t in enumerate(treatments, start=2):
                from app.config.constants import CANCER_TYPES, TREATMENT_STATUS, get_label
                
                ws_treatment.cell(row=row_idx, column=1, value=t["medical_id"])
                ws_treatment.cell(row=row_idx, column=2, value=t["name"])
                ws_treatment.cell(row=row_idx, column=3, value={"M": "男", "F": "女"}.get(t["gender"], ""))
                ws_treatment.cell(row=row_idx, column=4, value=get_label(CANCER_TYPES, t["cancer_type"]))
                ws_treatment.cell(row=row_idx, column=5, value=t["treatment_name"])
                ws_treatment.cell(row=row_idx, column=6, value=str(t["treatment_start"])[:10] if t["treatment_start"] else "")
                ws_treatment.cell(row=row_idx, column=7, value=str(t["treatment_end"])[:10] if t["treatment_end"] else "")
                ws_treatment.cell(row=row_idx, column=8, value=t["baseline_weight"])
                ws_treatment.cell(row=row_idx, column=9, value=t["last_weight"])
                ws_treatment.cell(row=row_idx, column=10, value=t["change_rate"])
                ws_treatment.cell(row=row_idx, column=11, value=t["weight_count"])
                ws_treatment.cell(row=row_idx, column=12, value=str(t["last_measure_date"])[:10] if t["last_measure_date"] else "")
                ws_treatment.cell(row=row_idx, column=13, value=get_label(TREATMENT_STATUS, t["status"]))
            
            # 調整欄寬
            for col in range(1, 14):
                ws_treatment.column_dimensions[chr(64 + col)].width = 12
            
            # === 介入表 ===
            ws_intervention = wb.create_sheet("介入記錄")
            
            int_headers = ["病歷號", "姓名", "療程", "類型", "來源", 
                          "狀態", "原因", "執行日期", "建立日期"]
            
            for col, header in enumerate(int_headers, start=1):
                cell = ws_intervention.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            interventions = self.get_intervention_report(start_date, end_date)
            for row_idx, i in enumerate(interventions, start=2):
                type_label = {"sdm": "SDM", "nutrition": "營養師"}.get(i["type"], i["type"])
                source_label = {"auto": "系統觸發", "manual": "手動"}.get(i["source"], i["source"])
                status_label = {"pending": "待處理", "completed": "已完成", "skipped": "已略過"}.get(i["status"], i["status"])
                
                ws_intervention.cell(row=row_idx, column=1, value=i["medical_id"])
                ws_intervention.cell(row=row_idx, column=2, value=i["name"])
                ws_intervention.cell(row=row_idx, column=3, value=i["treatment_name"])
                ws_intervention.cell(row=row_idx, column=4, value=type_label)
                ws_intervention.cell(row=row_idx, column=5, value=source_label)
                ws_intervention.cell(row=row_idx, column=6, value=status_label)
                ws_intervention.cell(row=row_idx, column=7, value=i["trigger_reason"] or i["skip_reason"] or "")
                ws_intervention.cell(row=row_idx, column=8, value=str(i["executed_at"])[:10] if i["executed_at"] else "")
                ws_intervention.cell(row=row_idx, column=9, value=str(i["created_at"])[:10] if i["created_at"] else "")
            
            # 調整欄寬
            for col in range(1, 10):
                ws_intervention.column_dimensions[chr(64 + col)].width = 12
            
            wb.save(filepath)
            return True
            
        except Exception as e:
            print(f"Excel 匯出失敗: {e}")
            return False
